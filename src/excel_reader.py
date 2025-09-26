#!/usr/bin/env python3
"""
Module de lecture Excel sans pandas pour éviter les conflits de dépendances
"""
import openpyxl
from io import BytesIO

def read_excel_file(file_content):
    """
    Lit un fichier Excel et retourne les données sous forme de dictionnaire.
    Détection automatique de la ligne d'en-têtes (si ce n'est pas la première ligne).
    """
    try:
        workbook = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
        worksheet = workbook.active

        # Candidats d'en-têtes attendus
        header_candidates = {
            'Nom PC', 'Nom Pc', 'Nom pc', 'Nom ordinateur', 'Nom machine',
            'Salle', 'Localisation', 'Emplacement', 'Lieu'
        }

        # Trouver la meilleure ligne d'en-têtes dans les 30 premières lignes
        best_row_idx = 1
        best_score = -1
        max_scan = min(30, worksheet.max_row)
        for r in range(1, max_scan + 1):
            vals = [str(c) if c is not None else '' for c in [cell.value for cell in worksheet[r]]]
            normalized = [v.strip() for v in vals]
            score = sum(1 for v in normalized if v in header_candidates)
            if score > best_score and any(v for v in normalized):
                best_score = score
                best_row_idx = r

        # Lire les en-têtes depuis la ligne détectée
        headers = []
        for cell in worksheet[best_row_idx]:
            headers.append(cell.value if cell.value is not None else '')

        # Lire les données à partir de la ligne suivante
        data = []
        for row in worksheet.iter_rows(min_row=best_row_idx + 1, values_only=True):
            # Ignorer les lignes complètement vides
            if all(v is None or str(v).strip() == '' for v in row):
                continue
            row_data = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    row_data[headers[i]] = value if value is not None else ''
            data.append(row_data)

        return {
            'columns': headers,
            'data': data,
            'total_rows': len(data)
        }

    except Exception as e:
        raise Exception(f"Erreur lors de la lecture du fichier Excel: {str(e)}")

def is_empty_value(value):
    """Vérifie si une valeur est vide ou nulle"""
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    if str(value).lower() in ['nan', 'none', 'null']:
        return True
    return False

def convert_boolean_field(value):
    """Convertit une valeur O/N en boolean"""
    if is_empty_value(value):
        return None
    
    value_str = str(value).strip().upper()
    if value_str.startswith('O'):
        return True
    elif value_str.startswith('N'):
        return False
    else:
        return None

def clean_string_field(value):
    """Nettoie une valeur string"""
    if is_empty_value(value):
        return None
    return str(value).strip()
